import requests
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment
from concurrent.futures import ThreadPoolExecutor, as_completed
import json
import threading
import time
import os
from pprint import pprint
import warnings
from urllib3.exceptions import InsecureRequestWarning
from kladr_dict import KLADR_CODES
from config import TOKEN
from config import TEMPLATE_PATH

# ─── Справочник статусов внутри кода ───────────────────────────────────
STATUS_LOOKUP = {
    1: "Прием заявок",
    4: "Отменено",
    5: "Не состоялось",
    2: "Работа комиссии",
    3: "Завершено",
    7: "Исполняется",
    6: "Исполнение завершено",
    8: "Расторжение",
    0: "Неизвестно"
}

# ─── Справочник ФЗ ───────────────────────────────────────────────
FZ_LOOKUP = {
    0: "223-ФЗ",
    1: "44-ФЗ",
}

# ─── Справочник способов проведения ────────────────────────────────────
PLACINGWAY_LOOKUP = {
    0: "ИС",  1: "ОК",  2: "ОА",  3: "ЭФ",  4: "ЗК",  5: "ПО",
    6: "ЕП",  7: "ОКУ", 8: "ОКД", 9: "ЗКК",10: "ЗККУ",11: "ЗККД",
   12: "ЗА", 13: "ЗКБ",14: "ЗП",15: "ЭА",16: "ИСМ",17: "СЗ",18: "ИОС",
   19: "РЕД",20: "ПЕР",21: "КП",22: "ЗКЭФ",23: "ОКЭФ",24: "ЗПЭФ",
   25: "ОКУЭФ",26: "ОКДЭФ",27: "ЗЦ",28: "ГА",29: "ПП"
}

# инвертируем KLADR_CODES: из кода региона (первые две цифры) → название
REGION_LOOKUP = {int(code[:2]): name for name, code in KLADR_CODES.items()}

# ─── Отключаем HTTPS-спам ─────────────────────────────────────────────
warnings.simplefilter('ignore', InsecureRequestWarning)

BASE_URL = 'https://tenderplan.ru/api'
REPORTS_DIR   = os.path.join(os.path.dirname(__file__), "reports")

headers = {
    'Authorization': f'Bearer {TOKEN}',
    'Accept': 'application/json',
    'Content-Type':  'application/json'}

session = requests.Session()
session.headers.update(headers) 


MAX_REQUESTS_PER_10_SECONDS = 250
TIME_WINDOW = 10.0  # секунд

lock = threading.Lock()
request_timestamps = []

def wait_for_rate_limit():
    with lock:
        now = time.time()
        # Удаляем устаревшие отметки (старше 10 секунд)
        while request_timestamps and request_timestamps[0] < now - TIME_WINDOW:
            request_timestamps.pop(0)
        # Если достигли лимита, ждем, пока освободится окно
        if len(request_timestamps) >= MAX_REQUESTS_PER_10_SECONDS:
            sleep_time = TIME_WINDOW - (now - request_timestamps[0])
            if sleep_time > 0:
                time.sleep(sleep_time)
        # Регистрируем новый запрос
        request_timestamps.append(time.time())


def generate_report(key_id: str) -> str:
    """
    Генерирует Excel-отчёт по tenderplan-ключу key_id.
    Возвращает путь к сохранённому файлу.
    """
    os.makedirs(REPORTS_DIR, exist_ok=True)

        # 1) Получаем список всех тендеров с пагинацией
    # ─── 2. Пагинация по /api/tenders/getlist с page/size ────────────────
    all_tenders = []
    page = 0
    size = 50  # сколько тендеров за 1 запрос

    while True:
        resp = requests.get(
            f"{BASE_URL}/tenders/v2/getlist",
            headers=headers,
            params={
            'type': 0,
            'id': key_id,
            'statuses': [1],
            'page': page,
            'size': size
        },
            verify=False
        )
        resp.raise_for_status()
        raw_batch = resp.json().get('tenders', [])
        # вручную фильтруем, чтобы точно остались только "Подача заявок"
        now_ts = int(datetime.now().timestamp() * 1000)
        batch = [
            t for t in raw_batch
            if t.get("status") == 1 and
            (t.get("submissionCloseDateTime") or t.get("submissionCloseDate") or 0) > now_ts
        ]
        if not batch:
            break
        all_tenders.extend(batch)
        page += 1
        # если пришло меньше, чем запрошено — значит это последняя страница
        if len(raw_batch) < size:
            break

    print(f"Всего тендеров после пагинации: {len(all_tenders)}")

    seen = set()
    unique_tenders = []
    for t in all_tenders:
        tid = t.get('_id')
        if tid in seen:
            continue
        seen.add(tid)
        unique_tenders.append(t)
    all_tenders = unique_tenders
    print(f"После удаления дубликатов: {len(all_tenders)} тендеров")

    # 2) Для каждого preview делаем detail-запрос и сохраняем в новом списке
    def fetch_detail(preview):
        rel_id = preview["_id"]
        for attempt in range(5):
            try:
                wait_for_rate_limit()
                r = session.get(
                    f"{BASE_URL}/tenders/get",
                    params={'id': rel_id},
                    verify=False
                )
                if r.status_code == 429:
                    # При 429 ждем с экспоненциальной задержкой
                    time.sleep(1 * (attempt + 1))
                    continue
                r.raise_for_status()
                det = r.json()
                # если вам нужен исходный статус для lookup'а:
                det["_preview_status"] = preview.get("status", 0)
                return det
            except requests.HTTPError as e:
                print(f"HTTPError при загрузке тендера {rel_id}: {e}")
                if attempt == 4:
                    raise
                time.sleep(1 * (attempt + 1))
            except Exception as e:
                print(f"Ошибка при загрузке тендера {rel_id}: {e}")
                if attempt == 4:
                    raise
                time.sleep(1 * (attempt + 1))
        raise Exception(f"Не удалось получить данные тендера {rel_id} после 5 попыток")
    # загружаем детали в параллельных потоках
    detailed = []
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = [executor.submit(fetch_detail, t) for t in all_tenders]
        for fut in as_completed(futures):
            detailed.append(fut.result())

    print("Получено детальных моделей тендеров:", len(detailed))
    # найдём максимальное время публикации среди тех, что попали в отчёт
    max_pub = max((d.get("publicationDate", 0) for d in detailed), default=0)
    # >>>> ДОБАВЛЯЕМ ЛОГ ДЛЯ ВЫВОДА ДАТ ПУБЛИКАЦИИ ВСЕХ ТЕНДЕРОВ <<<<
    print("\nВремя публикации тендеров:")
    for det in detailed:
        dt = det.get("publicationDate")
        if dt:
            dt_obj = datetime.fromtimestamp(dt / 1000)
            print(f"{det.get('number', det.get('_id'))}: {dt_obj} ({dt})")
        else:
            print(f"{det.get('number', det.get('_id'))}: дата не указана")

    # ─── 3. Открываем шаблон и очищаем предыдущие строки ───────────────────
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.value = None
            if hasattr(cell, 'hyperlink'):
                cell.hyperlink = None

    # ─── 4. Заполняем Excel (H — сразу maxPrice из getlist) ───────────────
    for idx, det in enumerate(detailed, start=3):
        _id      = det.get("_id")
        num      = det.get("number", _id)
        order    = det.get("orderName", "")
        price    = det.get("maxPrice", "") 
        if price is None or price == "":
            ws[f'I{idx}'] = "не установлена"
        else:
            ws[f'I{idx}'] = price             
        pub_dt   = datetime.fromtimestamp(det.get("publicationDate",0)/1000) if det.get("publicationDate") else ""
        close_ts = det.get("submissionCloseDateTime") or det.get("submissionCloseDate", 0)
        close_dt = datetime.fromtimestamp(close_ts/1000) if close_ts else ""
        customers = det.get('customers', [])
        customer_name = customers[0]["name"] if customers else ""
        ws[f'Q{idx}'] = customer_name
        # ЭТП
        plat = det.get('platform', {})

        #app_url = f"https://tenderplan.ru/app?key={key_id}&tender={_id}"

        ws[f'C{idx}'] = num
        ws[f'B{idx}'] = order
        ws[f'H{idx}'] = price                         # <— здесь

        # F: «ЕИС» – гиперссылка на ЕИС
        eis_link = det.get("href", "")
        if eis_link:
            ws[f'G{idx}'] = "Ссылка на тендер"
            ws[f'G{idx}'].hyperlink = eis_link
        else:
            ws[f'G{idx}'] = ""

        ws[f'A{idx}'] = pub_dt

        ws[f'H{idx}'] = plat.get("name", "")
        ws[f'H{idx}'].hyperlink = plat.get("href", "")

        ws[f'E{idx}'] = STATUS_LOOKUP.get(det.get("status"), "")
            # 1. получаем код ФЗ
        fz_id = det.get("type")  
        fz = FZ_LOOKUP.get(fz_id, "")

        # 2. получаем shortName типа торгов, как у вас было
        placing_code = det.get("placingWay")
        placing = ""
        if isinstance(placing_code, int):
            placing = PLACINGWAY_LOOKUP.get(placing_code, "")
        elif str(placing_code).isdigit():
            placing = PLACINGWAY_LOOKUP.get(int(placing_code), "")

        # 3. записываем в ячейку "Тип торгов"
        ws[f'F{idx}'] = " ".join(str(part) for part in (fz, placing) if part)
        prov = det.get('guaranteeProv')
        if prov is None or prov == "":
            ws[f'O{idx}'] = "не указано"
        else:
            ws[f'O{idx}'] = prov
        # ─── ОКПД2 ────────────────────────────────────────────────────────
        okpd2 = det.get("okpd2", "")
        if isinstance(okpd2, list):
            first = okpd2[0] if okpd2 else ""
            if isinstance(first, dict):
                okpd2_code = first.get("code", "") or first.get("fv", "")  # в зависимости от структуры
            else:
                okpd2_code = str(first)
        else:
            okpd2_code = str(okpd2)
        ws[f'D{idx}'] = okpd2_code

        # Обеспечение контракта
        contract_guarantee = det.get('guaranteeContract')
        if contract_guarantee is None or contract_guarantee == 0:
            ws[f'K{idx}'] = "указано в документации"
        else:
            ws[f'K{idx}'] = contract_guarantee

        app_guarantee = det.get('guaranteeApp')
        if app_guarantee is None or app_guarantee == 0:
            ws[f'J{idx}'] = "не требуется"
        else:
            ws[f'J{idx}'] = app_guarantee

        ws[f'L{idx}'] = det.get("currency", "")
        ws[f'M{idx}'] = close_dt
        ws[f'M{idx}'].number_format = 'DD.MM.YYYY HH:MM'
        region_id = det.get("region")  # это целое число, например 23
        ws[f'P{idx}'] = REGION_LOOKUP.get(region_id, "")
        sum_ts = det.get("summingUpDateTime")  # миллисекунды с эпохи, например 1690000000000
        if sum_ts:
            sum_dt = datetime.fromtimestamp(sum_ts / 1000)
            ws[f'N{idx}'] = sum_dt
            ws[f'N{idx}'].number_format = 'DD.MM.YYYY HH:MM'
        else:
            # если нет даты/времени — выводим текст по документации
            ws[f'N{idx}'] = "В соответствии с документацией о закупке"
            # выставляем формат «текст», чтобы Excel не пытался разобрать фразу как дату
            ws[f'N{idx}'].number_format = '@'
        #Документы к закупке  
        #atts = det.get("attachments", [])
        #cell = ws[f'R{idx}']
        #if atts:
        #    lines = []
        #    for a in atts:
        #        url  = a.get("href", "")
         #       name = a.get("displayName", url)
        #        lines.append(f"{name}: {url}")
         #   cell.value = "\n".join(str(line) for line in lines if line is not None and line != "")
         #   cell.alignment = Alignment(wrap_text=True, vertical='top')
        #    cell.number_format = 'General'    # <-- важно!
        #else:
          #  cell.value = ""
        # ─── Контакты заказчика ────────────────────────────────────────────
        raw = det.get("json", "")
        try:
            nested = json.loads(raw) if raw else {}
        except json.JSONDecodeError:
            nested = {}

        contacts_fv = nested.get("2", {}).get("fv", {})
        lines = []

        # Организация
        org = contacts_fv.get("0", {}).get("fv", "")
        if org:
            lines.append(org)

        # Фактический / почтовый адрес
        fact = contacts_fv.get("1", {}).get("fv", "")
        post = contacts_fv.get("2", {}).get("fv", "")
        if fact or post:
            lines.append(fact or post)

        # Массив контактов: FIO, Phone, Email
        for entry in contacts_fv.get("3", {}).get("fv", {}).values():
            fn = entry.get("fn")
            fv = entry.get("fv", "")
            if not fv:
                continue
            if fn == "FIO":
                lines.append(f"Контактное лицо: {fv}")
            elif fn == "Phone":
                lines.append(f"Телефон: {fv}")
            elif fn == "Email":
                lines.append(f"E-mail: {fv}")

        # Записываем в ячейку (например, столбец V), включаем переносы
        cell = ws[f'R{idx}']
        cell.value = "\n".join(str(line) for line in lines if line is not None and line != "")
        cell.alignment = Alignment(wrap_text=True)

    # ─── 5. Сохраняем ───────────────────────────────────────────────────────
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"тендеры_{now}.xlsx"
    out_path = os.path.join(REPORTS_DIR, filename)
    wb.save(out_path)

    return out_path, max_pub